"""Export-header → item-code crosswalk (FR-101–104; V6 zero-orphan rule).

Reads the survey export exactly as shipped (Qualtrics multi-header dialect:
row 1 platform question codes, row 2 question text with the designed item
codes embedded, row 3 import metadata), hashes the raw frame immediately
after load (FR-101), recovers the column→item mapping from the row-2 text
(FR-103), and enforces the accounting rules:

- every contract item resolves to exactly one column — a declared item the
  export lacks is a structural mismatch (FR-104; AT-M05-3);
- an item code matching several columns, or a column embedding several item
  codes, is ambiguity — a hard failure naming the columns (AT-M05-1);
- every export column resolves to exactly ONE role — model item,
  demographic, consent, id, completion, attention check, metadata, or
  ignored item; an unaccounted column is an orphan and a hard failure
  (V6, FR-507; AT-M05-2).

Halts are sink-emitting and artifact-free (this runs before a run directory
exists). Reports name columns and codes — never respondent values
(standards §7).
"""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from burhan.core.artifacts.canonical import sha256_canonical
from burhan.core.artifacts.models import StudyConfig
from burhan.core.errors import IntegrityHalt, halt

ROLE_MODEL_ITEM = "model_item"
ROLE_DEMOGRAPHIC = "demographic"
ROLE_CONSENT = "consent"
ROLE_ID = "id"
ROLE_COMPLETION = "completion"
ROLE_ATTENTION = "attention_check"
ROLE_METADATA = "metadata"
ROLE_IGNORED = "ignored_item"


@dataclass(frozen=True)
class Crosswalk:
    """The resolved export accounting for one study contract."""

    source_file: str
    header_rows: int
    n_data_rows: int
    raw_frame_sha256: str
    column_to_item: dict[str, str]
    roles: dict[str, str]

    def to_payload(self) -> dict[str, Any]:
        """Canonical-serializable crosswalk content (no respondent values)."""
        return {
            "source_file": self.source_file,
            "header_rows": self.header_rows,
            "n_data_rows": self.n_data_rows,
            "raw_frame_sha256": self.raw_frame_sha256,
            "column_to_item": dict(self.column_to_item),
            "roles": dict(self.roles),
        }


def build_crosswalk(export_path: Path, config: StudyConfig) -> Crosswalk:
    """Load the export, hash it, and resolve the full column accounting."""
    declared_format = str(config.data.format)
    rows = _load_raw(export_path, declared_format)
    raw_frame_sha256 = sha256_canonical(rows)  # immediately after load (FR-101)
    header_rows = _resolve_header_rows(config, rows, export_path)

    if len(rows) < header_rows:
        halt(
            IntegrityHalt(
                "export has fewer rows than the declared header block",
                report={
                    "file": export_path.name,
                    "rows": len(rows),
                    "header_rows": header_rows,
                },
            )
        )
    expected_width = len(rows[0])
    ragged = [
        {"row": index + 1, "expected_width": expected_width, "actual_width": len(row)}
        for index, row in enumerate(rows[:header_rows])
        if len(row) != expected_width
    ]
    if ragged:
        halt(
            IntegrityHalt(
                "ragged header block: header rows must all have the same width (FR-101/104)",
                report={"file": export_path.name, "ragged_header_rows": ragged},
            )
        )
    codes = rows[0]
    texts = rows[1] if header_rows >= 2 else rows[0]
    data_rows = rows[header_rows:]

    duplicates = sorted({code for code in codes if codes.count(code) > 1})
    if duplicates:
        halt(
            IntegrityHalt(
                "export declares duplicate column codes",
                report={"file": export_path.name, "columns": duplicates},
            )
        )

    column_to_item = _match_items(codes, texts, config, export_path)
    roles = _account_roles(codes, column_to_item, config, texts, export_path)

    return Crosswalk(
        source_file=export_path.name,
        header_rows=header_rows,
        n_data_rows=len(data_rows),
        raw_frame_sha256=raw_frame_sha256,
        column_to_item=column_to_item,
        roles=roles,
    )


def _has_qualtrics_signature(rows: list[list[str]]) -> bool:
    """The Qualtrics multi-header dialect marks its third header row (index 2)
    with per-column ImportId metadata JSON — a deterministic 3-header signature."""
    if len(rows) < 3 or not rows[2]:
        return False
    for cell in rows[2]:
        try:
            obj = json.loads(cell)
        except (json.JSONDecodeError, ValueError):
            return False
        if not (isinstance(obj, dict) and "ImportId" in obj):
            return False
    return True


def _resolve_header_rows(config: StudyConfig, rows: list[list[str]], export_path: Path) -> int:
    """Establish the header-row count without ever silently assuming one for a
    multi-header export (FR-104; TC-18, PLAN v1 §2). Precedence: (1) the contract's
    declared count wins; (2) else a recognized dialect — export_dialect == "qualtrics"
    or the row-3 ImportId signature — resolves it to three; (3) else an unambiguous
    single-header frame (row 0 already carries every modeled item code) resolves to one,
    preserving generic-CSV behavior; (4) else the run halts rather than mis-reading a
    multi-header export as single-header."""
    if config.data.header_rows is not None:
        return config.data.header_rows
    if config.data.export_dialect == "qualtrics" or _has_qualtrics_signature(rows):
        return 3
    if _is_unambiguous_single_header(config, rows):
        return 1
    halt(
        IntegrityHalt(
            "cannot establish the export header structure: declare data.header_rows "
            "or a recognized export_dialect — a multi-header export must not be read "
            "as single-header by default (FR-104)",
            report={
                "file": export_path.name,
                "header_rows": None,
                "export_dialect": config.data.export_dialect,
            },
        )
    )


def _is_unambiguous_single_header(config: StudyConfig, rows: list[list[str]]) -> bool:
    """A frame is consistent with a single header row when row 0 already carries every
    declared modeled item code — as a literal column name or a whole-token embedding —
    so reading it as one header loses no codes (PLAN v1 §2 step 3). A multi-header export
    hides its item codes in row 1, so its row 0 resolves none of them: that stays
    ambiguous and halts. Detection anchors on the modeled items only; role columns are
    resolved (and any ambiguity caught) downstream once the header count is fixed."""
    row0 = rows[0] if rows else []
    return all(_resolve_column(item.code, row0, row0) for item in config.instrument.items)


def _load_raw(export_path: Path, declared_format: str) -> list[list[str]]:
    """Read every cell as a string, exactly as shipped."""
    suffix = export_path.suffix.lower().lstrip(".")
    if suffix != declared_format:
        halt(
            IntegrityHalt(
                "export file extension does not match the contract's declared format",
                report={"file": export_path.name, "declared_format": declared_format},
            )
        )
    if not export_path.is_file():
        halt(
            IntegrityHalt(
                "export file missing",
                report={"file": str(export_path)},
            )
        )
    if declared_format == "csv":
        try:
            with export_path.open(newline="", encoding="utf-8") as handle:
                return [[cell for cell in row] for row in csv.reader(handle)]
        except (OSError, UnicodeDecodeError, csv.Error) as exc:
            halt(
                IntegrityHalt(
                    "export csv unreadable",
                    report={"file": export_path.name, "error": str(exc)},
                )
            )
    try:
        # Untyped third-party edge (no stubs in the locked dependency set).
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        book = load_workbook(export_path, read_only=True, data_only=True)
        sheet = book.active
        rows = [
            ["" if cell is None else str(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        book.close()
    except Exception as exc:  # noqa: BLE001 — openpyxl raises many types; retyped (NFR-201)
        halt(
            IntegrityHalt(
                "export xlsx unreadable",
                report={"file": export_path.name, "error": str(exc)},
            )
        )
    return rows


def _embeds(token: str, text: str) -> bool:
    """Whole-token match of a declared code inside export header text (FR-103)."""
    return re.search(rf"(?<![A-Za-z0-9_]){re.escape(token)}(?![A-Za-z0-9_])", text) is not None


def _resolve_column(token: str, codes: list[str], texts: list[str]) -> list[str]:
    """Every export column a declared token resolves to — by literal row-0 identity
    (the column name IS the token) or by whole-token embedding in the code-bearing
    header row. Non-modeled roles may be declared either way; a token resolving to
    more than one column is ambiguous (the caller halts)."""
    hits: list[str] = []
    for column, text in zip(codes, texts, strict=True):
        if column == token or _embeds(token, text):
            hits.append(column)
    return hits


def _match_items(
    codes: list[str], texts: list[str], config: StudyConfig, export_path: Path
) -> dict[str, str]:
    """Recover the column→item mapping from embedded row-2 item codes (FR-103)."""
    item_codes = [item.code for item in config.instrument.items]
    item_to_columns: dict[str, list[str]] = {code: [] for code in item_codes}
    column_to_items: dict[str, list[str]] = {}
    for column, text in zip(codes, texts, strict=True):
        for code in item_codes:
            if _embeds(code, text):
                item_to_columns[code].append(column)
                column_to_items.setdefault(column, []).append(code)

    ambiguous_items = {
        code: columns for code, columns in item_to_columns.items() if len(columns) > 1
    }
    multi_code_columns = {
        column: found for column, found in column_to_items.items() if len(found) > 1
    }
    if ambiguous_items or multi_code_columns:
        halt(
            IntegrityHalt(
                "ambiguous item-code embedding: crosswalk cannot be resolved "
                "without guessing (FR-103/104)",
                report={
                    "file": export_path.name,
                    "ambiguous_items": {k: v for k, v in sorted(ambiguous_items.items())},
                    "multi_code_columns": {k: v for k, v in sorted(multi_code_columns.items())},
                },
            )
        )
    missing = sorted(code for code, columns in item_to_columns.items() if not columns)
    if missing:
        halt(
            IntegrityHalt(
                "structural mismatch: contract declares items the export lacks (FR-104)",
                report={"file": export_path.name, "missing_items": missing},
            )
        )
    return {columns[0]: code for code, columns in item_to_columns.items()}


def _account_roles(
    codes: list[str],
    column_to_item: dict[str, str],
    config: StudyConfig,
    texts: list[str],
    export_path: Path,
) -> dict[str, str]:
    """Assign exactly one role to every export column (V6; FR-507). Declared
    non-modeled columns resolve like item codes — by literal row-0 name OR embedded
    header code — so a multi-header export need not name them as literal row-0
    identifiers (TC-18). Ambiguity and orphans remain hard failures."""
    roles: dict[str, str] = {}

    def claim(column: str, role: str) -> None:
        if column in roles:
            halt(
                IntegrityHalt(
                    "column claimed by two roles; accounting requires exactly one (V6)",
                    report={
                        "file": export_path.name,
                        "column": column,
                        "roles": sorted({roles[column], role}),
                    },
                )
            )
        roles[column] = role

    def resolve(token: str, role: str) -> None:
        columns = _resolve_column(token, codes, texts)
        if not columns:
            halt(
                IntegrityHalt(
                    "contract declares a column the export lacks (FR-104)",
                    report={"file": export_path.name, "column": token, "role": role},
                )
            )
        if len(columns) > 1:
            halt(
                IntegrityHalt(
                    "declared column resolves to more than one export column (FR-103/104)",
                    report={
                        "file": export_path.name,
                        "column": token,
                        "role": role,
                        "columns": sorted(columns),
                    },
                )
            )
        claim(columns[0], role)

    for column in column_to_item:  # already resolved to row-0 identifiers
        claim(column, ROLE_MODEL_ITEM)
    data = config.data
    if data.id_column is not None:
        resolve(data.id_column, ROLE_ID)
    if data.consent_column is not None:
        resolve(data.consent_column, ROLE_CONSENT)
    if data.completion is not None:
        if data.completion.progress_column is not None:
            resolve(data.completion.progress_column, ROLE_COMPLETION)
        if data.completion.finished_column is not None:
            resolve(data.completion.finished_column, ROLE_COMPLETION)
    for check in data.attention_checks or []:
        resolve(check.column, ROLE_ATTENTION)
    for demographic in data.demographics or []:
        resolve(demographic.column_hint, ROLE_DEMOGRAPHIC)
    for column in data.metadata_columns or []:
        resolve(column, ROLE_METADATA)
    for column in data.ignored_item_columns or []:
        resolve(column, ROLE_IGNORED)

    orphans = sorted(set(codes) - set(roles))
    if orphans:
        halt(
            IntegrityHalt(
                "unaccounted export columns: the zero-orphan rule is a hard failure (V6, FR-507)",
                report={"file": export_path.name, "orphans": orphans},
            )
        )
    return roles
